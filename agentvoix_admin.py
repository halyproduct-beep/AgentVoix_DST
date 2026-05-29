"""
AgentVoix — Plateforme Administration Complète
CIUSSS de l'Est-de-l'Île-de-Montréal — HMR
Version 3.0 — Toutes fonctions centralisées

Pages :
  🏠  Accueil          — KPIs temps réel + urgences
  📋  Requêtes         — tableau coordonnateur + case GMAO
  📞  Simulation       — démo appel complet DESS
  📊  Statistiques     — analytics Plotly
  🗂️  Documents        — gestion fichiers Excel
  🤖  Agent IA         — config OpenAI + test
  🔄  Workflow n8n     — statut + monitoring
  ⚙️  Configuration    — récepteurs, routage, sites
  👥  Utilisateurs     — gestion accès
  🔐  Connexion        — login 3 niveaux
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import os, json, hashlib, time

# ════════════════════════════════════════════════════════════════
# CONFIG PAGE
# ════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="AgentVoix Admin · HMR",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ════════════════════════════════════════════════════════════════
# CONSTANTES
# ════════════════════════════════════════════════════════════════
FICHIER_REGISTRE  = "Requetes_Téléphoniques_GMAO.xlsx"
FICHIER_LOCAUX    = "locaux_HMR.xls"
FICHIER_TACHES    = "Tâches_standards_CEMTL.xlsx"
FICHIER_INTERAL   = "Écran_web_V2__Intéral_.xlsx"
FICHIER_GMAO_LOG  = "gmao_transferts.json"
FICHIER_CONFIG    = "agentvoix_config.json"
FICHIER_USERS     = "agentvoix_users.json"

COLS_REGISTRE = [
    "N°", "Date d'ouverture", "Récepteur", "Émetteur",
    "N° Employé", "Demandé par", "Demandé pour",
    "Équipement / Local", "Usine / Site", "Pavillon",
    "Étage", "Aile", "Localisation", "Tél. #Ext.",
    "Remarque courte", "Remarques"
]

RECEPTEURS = {
    "HMR-CEPLOM":    ("Plomberie",         "🔧", "#1565C0"),
    "HMR-CEMAINT":   ("Maintenance",       "🔩", "#37474F"),
    "HMR-CEELE":     ("Électricité",       "⚡", "#F57F17"),
    "HMR-CEELEC":    ("Électromécanique",  "⚡", "#F57F17"),
    "HMR-CEMELE":    ("Électricité",       "⚡", "#F57F17"),
    "HMR-CEMELEC":   ("Électromécanique",  "⚡", "#E65100"),
    "HMR-CESER":     ("Serrurerie",        "🔑", "#6A1B9A"),
    "HMR-CEMMF":     ("Ventilation",       "💨", "#00695C"),
    "HMR-CEFRIG":    ("Frigoriste",        "❄️", "#0277BD"),
    "HMR-CEMENU":    ("Menuiserie",        "🪚", "#4E342E"),
    "HMR-CEMECA":    ("Mécanique",         "⚙️", "#455A64"),
    "HMR-CESALU":    ("Salubrité",         "🧹", "#2E7D32"),
    "HMR-CEOUVSPEC": ("Entretien général", "🏗️", "#546E7A"),
    "HMR-CEST":      ("Sous-traitant",     "📋", "#78909C"),
    "HMR-CEMELEC":   ("Électromécanique",  "⚡", "#E65100"),
}

PRIORITES = {
    "P1": ("Urgence < 1h",      "#C62828", "#FFEBEE"),
    "P2": ("Prioritaire < 4h",  "#E65100", "#FFF3E0"),
    "P3": ("Normal / jour",     "#1565C0", "#E3F2FD"),
    "P4": ("Différé / semaine", "#6A1B9A", "#F3E5F5"),
}

MOTS_URGENCE = [
    "urgence","urgent","danger","dangereux","feu","incendie","fumée",
    "explosion","blessé","accident","inondation","court-circuit",
    "électrocution","gaz","critique","patient en danger"
]

NON_SAIT_PAS = [
    "non","je sais pas","je ne sais pas","aucune idée","inconnu",
    "pas certain","nan","n/a","aucun","aucune","passer","skip",
]

QUESTIONS_AGENT = [
    ("demande_par",     "Quel est votre prénom et nom ?",                          "ex: Marie Tremblay"),
    ("telephone",       "Votre numéro de téléphone ou extension ?",                "ex: 4388385387 ou poste 1234"),
    ("no_employe",      "Votre numéro d'employé ?",                                "ex: HMR-8301"),
    ("demande_pour",    "Cette demande est pour vous-même ou une autre personne ?", "Dites 'moi' ou donnez le nom"),
    ("pavillon",        "Dans quel pavillon se trouve le problème ?",               "ex: Maisonneuve, Lavoisier, Guy-Bernier"),
    ("etage",           "À quel étage ?",                                           "RC, SS, ou un chiffre"),
    ("aile",            "Dans quelle aile ? (dites 'je ne sais pas' si inconnu)",   "ex: Aile A, B, C"),
    ("local",           "Quel est le numéro du local ou de la salle ?",             "ex: MA05102, LV01118"),
    ("localisation",    "Précisez la localisation exacte. (optionnel)",             "ex: Corridor nord, Salle de bain"),
    ("remarque_courte", "En quelques mots, quel est le problème ?",                 "ex: Fuite d'eau, Lumière brisée"),
    ("remarques",       "Décrivez le problème en détail.",                          "Plus de détails = intervention plus rapide"),
]

USERS_DEFAUT = {
    "admin":        {"mdp": hashlib.sha256("admin2026".encode()).hexdigest(),  "role": "admin",        "nom": "Administrateur"},
    "coord":        {"mdp": hashlib.sha256("coord2026".encode()).hexdigest(),  "role": "coordonnateur","nom": "Coordonnateur HMR"},
    "tech":         {"mdp": hashlib.sha256("tech2026".encode()).hexdigest(),   "role": "technicien",   "nom": "Technicien HMR"},
}

PAGES_PAR_ROLE = {
    "admin":         ["🏠 Accueil","📋 Requêtes","📞 Simulation","📊 Statistiques","🗂️ Documents","🤖 Agent IA","🎙️ ElevenLabs","🔄 Workflow n8n","⚙️ Configuration","👥 Utilisateurs"],
    "coordonnateur": ["🏠 Accueil","📋 Requêtes","📊 Statistiques","🗂️ Documents"],
    "technicien":    ["🏠 Accueil","📋 Requêtes"],
}

# ════════════════════════════════════════════════════════════════
# CSS
# ════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500&family=IBM+Plex+Sans:wght@300;400;500;600&display=swap');

html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; background:#F2F5F9; }

.top-bar {
    background: linear-gradient(135deg, #003A70 0%, #005FAD 60%, #00A1DE 100%);
    color: white; padding: 1rem 1.6rem; border-radius: 12px;
    margin-bottom: 1.2rem; display: flex;
    align-items: center; justify-content: space-between;
    box-shadow: 0 4px 20px rgba(0,58,112,0.25);
    position: relative; overflow: hidden;
}
.top-bar::after {
    content: \'\'; position: absolute; bottom: 0; left: 0; right: 0; height: 3px;
    background: linear-gradient(90deg, #EC1C24 0%, #F7941D 25%, #8DC63F 50%, #00A1DE 75%, #003A70 100%);
}
.top-bar h1 { font-size: 1.1rem; font-weight: 600; margin: 0; }
.top-bar .sub { font-size: 0.7rem; opacity: 0.7; margin-top: 2px; }
.top-bar .right { font-family: \'IBM Plex Mono\',monospace; font-size: 0.75rem; opacity: 0.8; text-align:right; }

.kpi { background: white; border: 1px solid #DDE3EC; border-radius: 12px;
    padding: 0.9rem 1rem; text-align: center;
    border-top: 3px solid #00A1DE;
    box-shadow: 0 2px 8px rgba(0,0,0,0.05); }
.kpi-n { font-size: 1.9rem; font-weight: 600; line-height: 1.1; color: #003A70; }
.kpi-l { font-size: 0.65rem; color: #6B7A8D; text-transform: uppercase;
    letter-spacing: 0.07em; margin-top: 4px; }

.req { background: white; border: 1px solid #DDE3EC; border-radius: 10px;
    padding: 0.85rem 1rem; margin-bottom: 0.55rem;
    transition: box-shadow 0.15s, transform 0.1s; }
.req:hover { box-shadow: 0 4px 16px rgba(0,58,112,0.1); transform: translateY(-1px); }
.req-urgent { border-left: 4px solid #EC1C24; }
.req-done { opacity: 0.55; }
.req-no { font-family:\'IBM Plex Mono\',monospace; font-size:0.73rem; font-weight:600; color:#003A70; }
.req-desc { font-size:0.8rem; background:#EBF6FB; border-left:3px solid #00A1DE;
    border-radius:0 6px 6px 0; padding:5px 9px; margin-top:5px; font-style:italic; color:#2A4A5E; }

.badge { display:inline-block; padding:2px 9px; border-radius:20px;
    font-size:0.67rem; font-weight:600; margin-right:3px; }

.msg-agent { background:#EBF6FB; border-left:3px solid #00A1DE; border-radius:0 10px 10px 10px;
    padding:9px 13px; margin:5px 0; max-width:78%; font-size:0.86rem; }
.msg-user { background:#003A70; color:white; border-radius:10px 0 10px 10px;
    padding:9px 13px; margin:5px 0 5px auto; max-width:68%; font-size:0.86rem; }
.msg-sys { background:#FFF8E1; border:1px solid #FFD54F; border-radius:7px;
    padding:7px 11px; margin:3px 0; font-size:0.75rem; color:#795548; }

.section-title { font-size:1rem; font-weight:600; color:#003A70;
    border-bottom:2px solid #00A1DE; padding-bottom:6px; margin-bottom:14px; }

[data-testid="stSidebar"] { background:#2C2F33 !important; border-right:1px solid #1a1d20; }
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] div { color:rgba(255,255,255,0.88) !important; }
[data-testid="stSidebar"] button {
    background: rgba(0,161,222,0.15) !important;
    border: 1px solid rgba(0,161,222,0.3) !important;
    color: white !important; border-radius: 8px !important;
}
[data-testid="stSidebar"] button:hover { background: rgba(0,161,222,0.3) !important; }

.config-card { background:white; border:1px solid #DDE3EC; border-radius:9px;
    padding:1rem; margin-bottom:0.6rem; border-left:3px solid #00A1DE; }
.config-title { font-size:0.88rem; font-weight:600; color:#003A70; margin-bottom:6px; }

.workflow-node { background:white; border:1px solid #DDE3EC; border-radius:8px;
    padding:0.7rem 0.9rem; margin-bottom:0.4rem; display:flex;
    align-items:center; gap:10px; }
.workflow-node:hover { box-shadow: 0 2px 8px rgba(0,58,112,0.08); }

.block-container { padding-top:0.7rem; }
.stApp { background: #F2F5F9 !important; }
#MainMenu, footer { visibility:hidden; }
hr { border-color:#DDE3EC; margin:0.7rem 0; }

.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #003A70, #00A1DE) !important;
    border: none !important; border-radius: 8px !important; font-weight: 500 !important;
}
.demo-box { margin-top:1.5rem; background:#EBF6FB; border-radius:8px;
    padding:0.8rem 1rem; font-size:0.72rem; color:#2A4A5E;
    line-height:1.8; border-left:3px solid #00A1DE; }
</style>
""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════
# UTILITAIRES
# ════════════════════════════════════════════════════════════════
def charger_json(fichier, defaut={}):
    if os.path.exists(fichier):
        try:
            with open(fichier, "r", encoding="utf-8") as f:
                return json.load(f)
        except: pass
    return defaut.copy()

def sauvegarder_json(fichier, data):
    with open(fichier, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def hash_mdp(mdp):
    return hashlib.sha256(mdp.encode()).hexdigest()

def badge(texte, fond, couleur):
    return f"<span class='badge' style='background:{fond};color:{couleur};'>{texte}</span>"

def est_urgent(row):
    t = (str(row.get("Remarque courte","")) + " " + str(row.get("Remarques",""))).lower()
    return any(m in t for m in MOTS_URGENCE)

def detecter_prio(rec, remarque=""):
    t = str(remarque).lower()
    if any(m in t for m in MOTS_URGENCE): return "P1"
    if rec in ("HMR-CEPLOM","HMR-CEELE","HMR-CEELEC","HMR-CEMMF","HMR-CEFRIG"): return "P2"
    if rec in ("HMR-CESER","HMR-CEMELE","HMR-CEMELEC","HMR-CEMECA"): return "P3"
    return "P3"

def classifier_recepteur(rc, rem):
    t = (rc + " " + rem).lower()
    if any(m in t for m in ["fuite","eau","plomb","toilette","bouché","drain","robinet","lavabo","douche"]): return "HMR-CEPLOM","Plomberie"
    if any(m in t for m in ["lumière","électr","prise","panne","interrupteur","ampoule","fluorescent","courant"]): return "HMR-CEELE","Électricité"
    if any(m in t for m in ["température","chaleur","froid","ventilation","climatisation","chauffage","thermostat","humidité"]): return "HMR-CEMMF","Ventilation"
    if any(m in t for m in ["serrure","porte","clé","verrou","poignée","cadenas","barre de panique"]): return "HMR-CESER","Serrurerie"
    if any(m in t for m in ["frigo","congélateur","réfrigérateur","machine à glace","climatiseur"]): return "HMR-CEFRIG","Frigoriste"
    if any(m in t for m in ["lit","civière","chariot","roulette","fauteuil","lève-personne","ridelle","manivelle"]): return "HMR-CEMAINT","Équip. soins"
    if any(m in t for m in ["meuble","plancher","plafond","tuile","fenêtre","mur","store","rideau"]): return "HMR-CEMENU","Menuiserie"
    if any(m in t for m in ["lavage","nettoyage","désinfection","poubelle","déchets","salubrité","souris","punaises"]): return "HMR-CESALU","Salubrité"
    return "HMR-CEMAINT","Maintenance"

@st.cache_data(ttl=20)
def charger_registre():
    if not os.path.exists(FICHIER_REGISTRE):
        return pd.DataFrame(columns=COLS_REGISTRE)
    try:
        wb = load_workbook(FICHIER_REGISTRE, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        hi = next((i for i, r in enumerate(rows)
                   if r[0] is not None and str(r[0]).startswith("N°")), None)
        if hi is None:
            return pd.DataFrame(columns=COLS_REGISTRE)
        data = [r for r in rows[hi+1:] if any(v is not None for v in r)]
        df = pd.DataFrame(data, columns=COLS_REGISTRE)
        col_no = next((c for c in df.columns if str(c).startswith("N°")), "N°")
        if col_no != "N°":
            df = df.rename(columns={col_no: "N°"})
        df = df[df["N°"].astype(str).str.startswith("BT-")]
        df["Date d'ouverture"] = pd.to_datetime(df["Date d'ouverture"], errors="coerce")
        for c in COLS_REGISTRE:
            if c in df.columns:
                df[c] = df[c].astype(str).str.strip().replace("nan","NAN")
        df["_urgent"]   = df.apply(est_urgent, axis=1)
        df["_priorite"] = df.apply(lambda r: "P1" if r["_urgent"] else detecter_prio(r["Récepteur"], r["Remarque courte"]), axis=1)
        df["_metier"]   = df["Récepteur"].apply(lambda r: RECEPTEURS.get(r, ("Autre","📋","#78909C"))[0])
        df["_ico"]      = df["Récepteur"].apply(lambda r: RECEPTEURS.get(r, ("Autre","📋","#78909C"))[1])
        df["_date_str"] = df["Date d'ouverture"].apply(lambda d: d.strftime("%Y-%m-%d %H:%M") if pd.notna(d) else "—")
        return df.reset_index(drop=True)
    except Exception as e:
        st.error(f"Erreur lecture registre : {e}")
        return pd.DataFrame(columns=COLS_REGISTRE)

@st.cache_data(ttl=60)
def charger_taches():
    try:
        df = pd.read_excel(FICHIER_TACHES, header=1)
        df = df.rename(columns={
            "N° tâche standard":"no_tache","Description":"description",
            "Catégorie":"categorie","Récepteur":"recepteur","Ressource":"ressource","Usine":"usine"
        })
        df = df[df["no_tache"].astype(str).str.startswith("HMR")].dropna(subset=["no_tache"])
        return df[["no_tache","description","categorie","recepteur","ressource"]].reset_index(drop=True)
    except Exception as e:
        return pd.DataFrame()

@st.cache_data(ttl=300)
def charger_locaux(n=100):
    try:
        df = pd.read_excel(FICHIER_LOCAUX, engine="xlrd", nrows=n)
        df = df[["N° Équipement/Local","Description","Chemin d'accès"]].dropna(subset=["N° Équipement/Local"])
        df.columns = ["numero","description","chemin"]
        return df
    except:
        return pd.DataFrame()

def sauvegarder_bt_excel(bon):
    bord = Border(left=Side(style="thin"),right=Side(style="thin"),
                  top=Side(style="thin"),bottom=Side(style="thin"))
    al = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ac = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ligne = [
        bon["no_bon"], bon["date"], bon["recepteur"], "Agent IA",
        bon["no_employe"], bon["demande_par"], bon["demande_pour"],
        bon["local"], "HMR", bon["pavillon"], bon["etage"],
        bon["aile"], bon["localisation"], bon["telephone"],
        bon["remarque_courte"], bon["remarques"],
    ]
    if os.path.exists(FICHIER_REGISTRE):
        wb = load_workbook(FICHIER_REGISTRE)
        ws = wb.active
    else:
        wb = Workbook(); ws = wb.active; ws.title = "Registre"
        ws.merge_cells("A1:P1")
        ws["A1"].value = "REGISTRE DES APPELS DE SERVICE – GMAO"
        ws["A1"].font  = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill  = PatternFill("solid", fgColor="0D2B55")
        ws["A1"].alignment = ac
        ws.merge_cells("A2:P2")
        ws["A2"].value = f"Généré le : {datetime.now().strftime('%Y-%m-%d')}"
        ws["A2"].font  = Font(italic=True, size=9, color="5B6E7C")
        for col, h in enumerate(COLS_REGISTRE, 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = Font(bold=True, size=9, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="006D77")
            c.alignment = ac; c.border = bord
        for i,w in enumerate([18,20,16,12,14,22,22,20,12,24,8,10,28,14,25,40],1):
            ws.column_dimensions[ws.cell(1,i).column_letter].width = w
    next_row = max(ws.max_row + 1, 4)
    for col, val in enumerate(ligne, 1):
        c = ws.cell(row=next_row, column=col, value=val)
        c.border = bord; c.alignment = al
        if col <= 4:
            c.fill = PatternFill("solid", fgColor="EDF6F9")
            c.font = Font(size=9, bold=True, color="006D77")
        else:
            c.fill = PatternFill("solid", fgColor="FFFFFF")
            c.font = Font(size=9)
        if str(val) == "NAN":
            c.font = Font(size=9, italic=True, color="AAAAAA")
        if bon.get("urgence") and col == 1:
            c.fill = PatternFill("solid", fgColor="C0392B")
            c.font = Font(size=9, bold=True, color="FFFFFF")
    ws.freeze_panes = "A4"
    try:
        wb.save(FICHIER_REGISTRE)
        return True
    except Exception as e:
        st.error(f"Erreur sauvegarde : {e}")
        return False

# ════════════════════════════════════════════════════════════════
# AUTHENTIFICATION
# ════════════════════════════════════════════════════════════════
def init_auth():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
        st.session_state.user      = None
        st.session_state.role      = None
        st.session_state.nom       = None

def page_login():
    # Logo CIUSSS en SVG + design professionnel pleine page
    st.markdown("""
    <style>
    .stApp { background: #f0f4f8; }
    .login-page {
        display: flex; min-height: 100vh;
        align-items: center; justify-content: center;
        padding: 2rem;
    }
    .login-wrapper {
        display: flex; width: 100%; max-width: 900px;
        box-shadow: 0 20px 60px rgba(0,0,0,0.12);
        border-radius: 16px; overflow: hidden;
    }
    .login-left {
        flex: 1; background: linear-gradient(145deg, #003A70 0%, #0066B3 50%, #00A1DE 100%);
        padding: 3rem 2.5rem; display: flex; flex-direction: column;
        justify-content: space-between; min-height: 520px;
    }
    .login-right {
        flex: 1; background: white;
        padding: 3rem 2.5rem; display: flex;
        flex-direction: column; justify-content: center;
    }
    .ciusss-logo-text {
        font-size: 1.4rem; font-weight: 700; color: white;
        line-height: 1.3; margin-bottom: 0.5rem;
    }
    .ciusss-sub {
        font-size: 0.78rem; color: rgba(255,255,255,0.75);
        line-height: 1.5;
    }
    .login-tagline {
        font-size: 0.85rem; color: rgba(255,255,255,0.9);
        line-height: 1.7; border-top: 1px solid rgba(255,255,255,0.2);
        padding-top: 1.5rem; margin-top: 1.5rem;
    }
    .login-right-title {
        font-size: 1.5rem; font-weight: 700; color: #003A70;
        margin-bottom: 0.3rem;
    }
    .login-right-sub {
        font-size: 0.8rem; color: #8a9ab0; margin-bottom: 2rem;
    }
    .demo-box {
        margin-top: 1.5rem; background: #f4f8fc;
        border-radius: 8px; padding: 0.8rem 1rem;
        font-size: 0.72rem; color: #7a8a9a; line-height: 1.8;
        border-left: 3px solid #0066B3;
    }
    </style>

    <div class="login-page">
      <div class="login-wrapper">

        <!-- Côté gauche : identité CIUSSS -->
        <div class="login-left">
          <div>
            <!-- Logo CIUSSS SVG -->
            <svg width="80" height="80" viewBox="0 0 120 120" style="margin-bottom:1.2rem">
              <circle cx="60" cy="60" r="58" fill="none" stroke="rgba(255,255,255,0.15)" stroke-width="2"/>
              <!-- Arcs colorés CIUSSS -->
              <path d="M60 15 A45 45 0 0 1 95 38" fill="none" stroke="#00C4B4" stroke-width="10" stroke-linecap="round"/>
              <path d="M98 45 A45 45 0 0 1 90 92" fill="none" stroke="#F7941D" stroke-width="10" stroke-linecap="round"/>
              <path d="M83 99 A45 45 0 0 1 37 99" fill="none" stroke="#8DC63F" stroke-width="10" stroke-linecap="round"/>
              <path d="M28 92 A45 45 0 0 1 20 45" fill="none" stroke="#00A1DE" stroke-width="10" stroke-linecap="round"/>
              <path d="M23 38 A45 45 0 0 1 58 15" fill="none" stroke="#EC1C24" stroke-width="10" stroke-linecap="round"/>
              <!-- Point central -->
              <circle cx="60" cy="60" r="8" fill="white" opacity="0.9"/>
            </svg>
            <div class="ciusss-logo-text">CIUSSS de l'Est-de-<br>l'Île-de-Montréal</div>
            <div class="ciusss-sub">Direction des services techniques<br>Hôpital Maisonneuve-Rosemont</div>
          </div>

          <div>
            <div style="font-size:1.8rem;font-weight:700;color:white;margin-bottom:0.5rem">
              AgentVoix
            </div>
            <div style="font-size:0.82rem;color:rgba(255,255,255,0.8);line-height:1.6">
              Système intelligent de gestion<br>des appels de service technique
            </div>
            <div class="login-tagline">
              ✦ Disponible 24h/7j<br>
              ✦ Conformité Loi 25<br>
              ✦ Intégration GMAO Intéral<br>
              ✦ Classification IA automatique
            </div>
          </div>
        </div>

        <!-- Côté droit : formulaire -->
        <div class="login-right">
          <div class="login-right-title">Connexion</div>
          <div class="login-right-sub">Plateforme d'administration AgentVoix — HMR</div>
        </div>

      </div>
    </div>
    """, unsafe_allow_html=True)

    # Formulaire Streamlit centré
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        with st.form("login_form"):
            user = st.text_input("Identifiant", placeholder="admin / coord / tech")
            mdp  = st.text_input("Mot de passe", type="password", placeholder="••••••••")
            submitted = st.form_submit_button(
                "🔐  Se connecter",
                use_container_width=True,
                type="primary"
            )

        if submitted:
            users = charger_json(FICHIER_USERS, USERS_DEFAUT)
            if user in users and users[user]["mdp"] == hash_mdp(mdp):
                st.session_state.logged_in = True
                st.session_state.user      = user
                st.session_state.role      = users[user]["role"]
                st.session_state.nom       = users[user]["nom"]
                st.rerun()
            else:
                st.error("❌ Identifiant ou mot de passe incorrect.")

        st.markdown("""
        <div class="demo-box">
        <b>Comptes de démonstration :</b><br>
        admin &nbsp;/ admin2026 &nbsp;→ Administrateur<br>
        coord / coord2026 → Coordonnateur<br>
        tech &nbsp;&nbsp;/ tech2026 &nbsp;&nbsp;→ Technicien
        </div>
        """, unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════
# SIDEBAR
# ════════════════════════════════════════════════════════════════
def render_sidebar(df_raw, gmao_log):
    with st.sidebar:
        st.markdown(f"""
        <div style='margin-bottom:1rem'>
          <div style='font-size:1rem;font-weight:600'>🏥 AgentVoix</div>
          <div style='font-size:0.72rem;opacity:0.65'>HMR · Maintenance des bâtiments</div>
          <div style='margin-top:8px;background:rgba(255,255,255,0.1);border-radius:6px;padding:6px 10px;font-size:0.75rem'>
            👤 {st.session_state.nom}<br>
            <span style='opacity:0.65'>{st.session_state.role.capitalize()}</span>
          </div>
        </div>
        """, unsafe_allow_html=True)

        pages_dispo = PAGES_PAR_ROLE.get(st.session_state.role, ["🏠 Accueil"])
        page = st.radio("", pages_dispo, label_visibility="collapsed")

        st.markdown("---")

        if "Requêtes" in page:
            st.markdown("**Filtres**")
            f_statut = st.selectbox("Statut", ["Toutes","Non transférées","Transférées"])
            f_prio   = st.selectbox("Priorité", ["Toutes","P1","P2","P3","P4"])
            f_metier = st.selectbox("Métier", ["Tous"] + sorted(set(v[0] for v in RECEPTEURS.values())))
            f_search = st.text_input("🔎 Recherche", placeholder="N°BT, nom, local...")
        else:
            f_statut = "Toutes"; f_prio = "Toutes"
            f_metier = "Tous";   f_search = ""

        st.markdown("---")

        col_ref, col_quit = st.columns(2)
        with col_ref:
            if st.button("🔄 Actualiser", use_container_width=True):
                st.cache_data.clear()
                st.rerun()
        with col_quit:
            if st.button("🚪 Déconnexion", use_container_width=True):
                for k in list(st.session_state.keys()):
                    del st.session_state[k]
                st.rerun()

        # Stats rapides sidebar
        if not df_raw.empty:
            total     = len(df_raw)
            urgences  = int(df_raw["_urgent"].sum())
            transferes = sum(1 for n in df_raw["N°"] if n in gmao_log)
            st.markdown(f"""
            <div style='font-size:0.68rem;margin-top:0.5rem;opacity:0.6;line-height:2'>
            📊 {total} requêtes · 🚨 {urgences} urgences<br>
            ✅ {transferes} transférées GMAO<br>
            AgentVoix v3.0 · {datetime.now().strftime('%H:%M')}
            </div>
            """, unsafe_allow_html=True)

    return page, f_statut, f_prio, f_metier, f_search

# ════════════════════════════════════════════════════════════════
# PAGE ACCUEIL
# ════════════════════════════════════════════════════════════════
def page_accueil(df_raw, gmao_log):
    total      = len(df_raw)
    urgences   = int(df_raw["_urgent"].sum()) if not df_raw.empty else 0
    transferes = sum(1 for n in df_raw["N°"] if n in gmao_log) if not df_raw.empty else 0
    attente    = total - transferes

    c1,c2,c3,c4 = st.columns(4)
    for col,(num,label,couleur) in zip([c1,c2,c3,c4],[
        (total,       "Requêtes totales",   "#0D2B55"),
        (attente,     "En attente GMAO",    "#E65100"),
        (urgences,    "Urgences actives",   "#C62828"),
        (transferes,  "Transférées GMAO",   "#2E7D32"),
    ]):
        with col:
            st.markdown(f"""
            <div class="kpi">
              <div class="kpi-n" style="color:{couleur};">{num}</div>
              <div class="kpi-l">{label}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    cl, cr = st.columns([3,2])

    with cl:
        st.markdown('<div class="section-title">🚨 Urgences et priorités P1-P2</div>', unsafe_allow_html=True)
        if df_raw.empty:
            st.info("Aucune requête dans le registre. Lance agentvoix pour recevoir des appels.")
        else:
            df_p = df_raw[df_raw["_priorite"].isin(["P1","P2"]) & ~df_raw["N°"].apply(lambda n: n in gmao_log)]
            if df_p.empty:
                st.success("✅ Aucune urgence en attente actuellement.")
            else:
                for _,row in df_p.iterrows():
                    p = row["_priorite"]; pl,pc,pf = PRIORITES[p]
                    metier,ico,_ = RECEPTEURS.get(row["Récepteur"],("Autre","📋","#78909C"))
                    st.markdown(f"""
                    <div class="req req-urgent">
                      <div style="display:flex;align-items:center;gap:7px;margin-bottom:5px">
                        <span class="req-no">{row['N°']}</span>
                        {badge(f"🚨 {p}",pf,pc)}
                        {badge(f"{ico} {metier}","#f0f0f0","#333")}
                        <span style="margin-left:auto;font-size:0.68rem;color:#aaa">{row['_date_str']}</span>
                      </div>
                      <div style="font-size:0.87rem;font-weight:500;color:#1a2a3a">{row['Remarque courte']}</div>
                      <div style="font-size:0.75rem;color:#8a9ab0;margin-top:2px">
                        📍 {row['Pavillon']} · {row['Équipement / Local']} &nbsp;|&nbsp;
                        👤 {row['Demandé par']} · {row['Tél. #Ext.']}
                      </div>
                    </div>""", unsafe_allow_html=True)

    with cr:
        st.markdown('<div class="section-title">📊 Requêtes par métier</div>', unsafe_allow_html=True)
        if not df_raw.empty:
            mc = df_raw["_metier"].value_counts().reset_index()
            mc.columns = ["Métier","N"]
            fig = px.bar(mc, x="N", y="Métier", orientation="h",
                         color="N", color_continuous_scale="Blues", height=260)
            fig.update_layout(margin=dict(l=0,r=0,t=5,b=0),
                              showlegend=False, coloraxis_showscale=False,
                              plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
            fig.update_traces(marker_line_width=0)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Aucune donnée.")

    st.markdown('<div class="section-title">⏱️ Activité récente</div>', unsafe_allow_html=True)
    if not df_raw.empty:
        for _,row in df_raw.sort_values("Date d'ouverture", ascending=False).head(5).iterrows():
            p = row["_priorite"]; pl,pc,pf = PRIORITES[p]
            done = row["N°"] in gmao_log
            st.markdown(f"""
            <div class="req {'req-done' if done else ''}">
              <div style="display:flex;align-items:center;gap:7px">
                <span class="req-no">{row['N°']}</span>
                {badge(p,pf,pc)}
                {badge("✅ GMAO" if done else "🤖 IA","#e8f5e9" if done else "#e3f2fd","#2e7d32" if done else "#1565c0")}
                <span style="font-size:0.85rem;font-weight:500;color:#1a2a3a">{row['Remarque courte']}</span>
                <span style="margin-left:auto;font-size:0.68rem;color:#aaa">{row['_date_str']}</span>
              </div>
            </div>""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════
# PAGE REQUÊTES
# ════════════════════════════════════════════════════════════════
def page_requetes(df_raw, gmao_log, f_statut, f_prio, f_metier, f_search):
    st.markdown('<div class="section-title">📋 Requêtes — Vue coordonnateur</div>', unsafe_allow_html=True)
    st.markdown("Cochez ✓ quand vous avez saisi la requête dans la **GMAO Intéral**.")

    df = df_raw.copy() if not df_raw.empty else df_raw
    if not df.empty:
        if f_statut == "Non transférées": df = df[~df["N°"].apply(lambda n: n in gmao_log)]
        elif f_statut == "Transférées":   df = df[df["N°"].apply(lambda n: n in gmao_log)]
        if f_prio != "Toutes":            df = df[df["_priorite"] == f_prio]
        if f_metier != "Tous":            df = df[df["_metier"] == f_metier]
        if f_search:
            s = f_search.lower()
            mask = (df["N°"].str.lower().str.contains(s,na=False) |
                    df["Demandé par"].str.lower().str.contains(s,na=False) |
                    df["Équipement / Local"].str.lower().str.contains(s,na=False) |
                    df["Remarque courte"].str.lower().str.contains(s,na=False))
            df = df[mask]

    if df.empty:
        st.info("Aucune requête. Lance `agentvoix_complet.py` pour recevoir des appels." if df_raw.empty
                else "Aucune requête pour ces filtres.")
        return

    st.markdown(f"**{len(df)} requête(s) affichée(s)**")
    df_sorted = df.sort_values(["_urgent","Date d'ouverture"], ascending=[False,False])

    for idx, row in df_sorted.iterrows():
        no_bt     = row["N°"]
        urgent    = row["_urgent"]
        transfere = no_bt in gmao_log
        p         = row["_priorite"]
        pl,pc,pf  = PRIORITES[p]
        metier,ico,mc = RECEPTEURS.get(row["Récepteur"],("Autre","📋","#78909C"))

        col_card, col_cb = st.columns([11,1])
        with col_card:
            st.markdown(f"""
            <div class="req {'req-urgent' if urgent else ''} {'req-done' if transfere else ''}">
              <div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;margin-bottom:5px">
                <span class="req-no">{no_bt}</span>
                {"<span class='badge' style='background:#FFEBEE;color:#C62828;'>🚨 URGENCE</span>" if urgent else ""}
                <span class="badge" style="background:{pf};color:{pc};">{p} — {pl}</span>
                {"<span class='badge' style='background:#e8f5e9;color:#2e7d32;'>✅ Transféré GMAO</span>" if transfere else ""}
                <span style="margin-left:auto;font-size:0.68rem;color:#aaa">{row['_date_str']}</span>
              </div>
              <div style="display:grid;grid-template-columns:repeat(5,1fr);gap:8px;margin-bottom:5px">
                <div><div style="font-size:0.62rem;color:#aaa;text-transform:uppercase">Métier</div>
                     <div style="font-size:0.83rem;font-weight:500">{ico} {metier}</div></div>
                <div><div style="font-size:0.62rem;color:#aaa;text-transform:uppercase">Récepteur</div>
                     <div style="font-size:0.78rem;font-family:'IBM Plex Mono',monospace">{row['Récepteur']}</div></div>
                <div><div style="font-size:0.62rem;color:#aaa;text-transform:uppercase">Pavillon / Local</div>
                     <div style="font-size:0.83rem;font-weight:500">{row['Pavillon']} · {row['Équipement / Local']}</div></div>
                <div><div style="font-size:0.62rem;color:#aaa;text-transform:uppercase">Demandé par</div>
                     <div style="font-size:0.83rem">{row['Demandé par']}</div></div>
                <div><div style="font-size:0.62rem;color:#aaa;text-transform:uppercase">Téléphone</div>
                     <div style="font-size:0.83rem">{row['Tél. #Ext.']}</div></div>
              </div>
              <div class="req-desc"><b>Problème :</b> {row['Remarque courte']}<br>
              <span style="font-size:0.74rem;color:#5a7080">{row['Remarques']}</span></div>
            </div>""", unsafe_allow_html=True)

        with col_cb:
            st.markdown("<div style='height:26px'></div>", unsafe_allow_html=True)
            checked = st.checkbox("✓", value=transfere, key=f"cb_{no_bt}_{idx}",
                                  disabled=transfere,
                                  help="Cocher = confirmé dans la GMAO Intéral")
            if checked and not transfere:
                gmao_log[no_bt] = {
                    "transfere_le": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "coordonnateur": st.session_state.nom,
                    "role": st.session_state.role,
                }
                sauvegarder_json(FICHIER_GMAO_LOG, gmao_log)
                st.cache_data.clear()
                st.rerun()

# ════════════════════════════════════════════════════════════════
# PAGE SIMULATION
# ════════════════════════════════════════════════════════════════
def page_simulation():
    st.markdown('<div class="section-title">📞 Simulation d\'appel — Démo AgentVoix</div>', unsafe_allow_html=True)
    st.markdown("Simulez un appel complet. Le bon de travail est généré et sauvegardé dans Excel.")

    for k,v in [("sim_step",-1),("sim_data",{}),("sim_history",[]),
                ("sim_done",False),("sim_bon",None)]:
        if k not in st.session_state: st.session_state[k] = v

    col_chat, col_form = st.columns([3,2])

    with col_chat:
        st.markdown("**💬 Conversation**")
        chat_html = ""
        for role,msg in st.session_state.sim_history:
            if role=="agent":   chat_html += f'<div class="msg-agent">🤖 <b>Agent IA :</b> {msg}</div>'
            elif role=="user":  chat_html += f'<div class="msg-user">👤 {msg}</div>'
            elif role=="system":chat_html += f'<div class="msg-sys">⚙️ {msg}</div>'
        if chat_html:
            st.markdown(f'<div style="background:#f8fafc;border:1px solid #e8edf3;border-radius:9px;padding:0.9rem;max-height:400px;overflow-y:auto">{chat_html}</div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

        intro = ("Bonjour ! Service de maintenance des bâtiments de l'Hôpital Maisonneuve-Rosemont. "
                 "Je suis l'agent IA. <b>Cet appel est traité par une intelligence artificielle</b> et sera "
                 "enregistré uniquement aux fins de traitement de votre demande (Loi 25). "
                 "Dites <em>'humain'</em> en tout temps pour parler à une personne. "
                 "Je vais maintenant enregistrer votre demande.")

        if st.session_state.sim_step == -1:
            st.markdown(f'<div class="msg-agent">🤖 <b>Agent IA :</b> {intro}</div>', unsafe_allow_html=True)
            if st.button("▶️ Démarrer l'appel", type="primary", use_container_width=True):
                st.session_state.sim_history.append(("agent", intro))
                st.session_state.sim_step = 0
                cle,q,a = QUESTIONS_AGENT[0]
                st.session_state.sim_history.append(("agent", f"{q} <em style='color:#8a9ab0;font-size:0.78rem'>({a})</em>"))
                st.session_state.sim_data["_derniere_cle"] = cle
                st.rerun()

        elif not st.session_state.sim_done:
            step = st.session_state.sim_step
            if step < len(QUESTIONS_AGENT):
                cle,q,a = QUESTIONS_AGENT[step]
                with st.form(key=f"sim_{step}", clear_on_submit=True):
                    rep = st.text_input("Votre réponse :", placeholder=f"{a}  |  'je ne sais pas' si inconnu")
                    sub = st.form_submit_button("Envoyer ↵", use_container_width=True)
                if sub and rep.strip():
                    rep = rep.strip()
                    if any(m in rep.lower() for m in MOTS_URGENCE):
                        st.session_state.sim_history.append(("user", rep))
                        st.session_state.sim_history.append(("system","🚨 URGENCE — Transfert vers permanence HMR !"))
                        st.session_state.sim_data[cle] = rep
                        st.session_state.sim_data["urgence"] = True
                        for q2,_,_ in QUESTIONS_AGENT:
                            if q2 not in st.session_state.sim_data:
                                st.session_state.sim_data[q2] = "NAN"
                        st.session_state.sim_step = len(QUESTIONS_AGENT)
                    else:
                        valeur = "NAN" if any(n in rep.lower() for n in NON_SAIT_PAS) else rep
                        st.session_state.sim_data[cle] = valeur
                        st.session_state.sim_history.append(("user", rep))
                        if valeur == "NAN":
                            st.session_state.sim_history.append(("agent","D'accord, j'inscris 'NAN' pour ce champ."))
                        if cle == "demande_pour" and rep.lower() in ("moi","moi-même","pour moi","même","pareil"):
                            st.session_state.sim_data["demande_pour"] = st.session_state.sim_data.get("demande_par","NAN")
                        st.session_state.sim_step += 1
                    ns = st.session_state.sim_step
                    if ns < len(QUESTIONS_AGENT):
                        c2,q2,a2 = QUESTIONS_AGENT[ns]
                        st.session_state.sim_history.append(("agent",f"{q2} <em style='color:#8a9ab0;font-size:0.78rem'>({a2})</em>"))
                    st.rerun()

            if st.session_state.sim_step >= len(QUESTIONS_AGENT) and not st.session_state.sim_done:
                d = st.session_state.sim_data
                rec,metier = classifier_recepteur(d.get("remarque_courte",""), d.get("remarques",""))
                urgent = d.get("urgence",False) or any(m in d.get("remarque_courte","").lower() for m in MOTS_URGENCE)
                prio   = "P1" if urgent else detecter_prio(rec, d.get("remarque_courte",""))
                now    = datetime.now()
                no_bt  = f"BT-HMR-{now.strftime('%Y%m%d')}-{str(int(now.timestamp()))[-4:]}"
                bon = {
                    "no_bon":now.strftime and no_bt, "date":now.strftime("%Y-%m-%d %H:%M:%S"),
                    "recepteur":rec, "no_employe":d.get("no_employe","NAN"),
                    "demande_par":d.get("demande_par","NAN"), "demande_pour":d.get("demande_pour","NAN"),
                    "pavillon":d.get("pavillon","NAN"), "etage":d.get("etage","NAN"),
                    "aile":d.get("aile","NAN"), "local":d.get("local","NAN"),
                    "localisation":d.get("localisation","NAN"), "telephone":d.get("telephone","NAN"),
                    "remarque_courte":d.get("remarque_courte","NAN"), "remarques":d.get("remarques","NAN"),
                    "metier":metier, "priorite":prio, "urgence":urgent,
                }
                if urgent:
                    msg_fin = f"🚨 URGENCE — Transfert immédiat. Numéro BT : <b>{no_bt}</b>. Ne raccrochez pas !"
                else:
                    msg_fin = (f"Merci {d.get('demande_par','')} ! Demande enregistrée sous <b>{no_bt}</b>. "
                               f"L'équipe {metier} va intervenir. Priorité {prio}. Bonne journée !")
                st.session_state.sim_history.append(("agent", msg_fin))
                st.session_state.sim_done = True
                st.session_state.sim_bon  = bon
                sauvegarder_bt_excel(bon)
                st.cache_data.clear()
                st.rerun()
        else:
            st.success("✅ Appel terminé — BT sauvegardé dans le registre Excel.")
            if st.button("🔄 Nouvel appel", use_container_width=True):
                for k in ["sim_step","sim_data","sim_history","sim_done","sim_bon"]:
                    if k in st.session_state: del st.session_state[k]
                st.rerun()

    with col_form:
        st.markdown("**📄 Bon de travail**")
        d   = st.session_state.sim_data
        bon = st.session_state.sim_bon
        src = bon if bon else d
        if not src:
            st.markdown('<div style="background:#f8fafc;border:1px dashed #cdd5de;border-radius:9px;padding:2rem;text-align:center;color:#aaa">Le bon apparaîtra ici au fur et à mesure.</div>', unsafe_allow_html=True)
        else:
            rec  = src.get("recepteur","—")
            prio = src.get("priorite","P3")
            mi   = RECEPTEURS.get(rec,("Autre","📋","#78909C"))
            pl,pc,pf = PRIORITES.get(prio,("—","gray","#eee"))
            champs = [
                ("N° BT",        src.get("no_bon","— en cours —")),
                ("Récepteur",    rec),
                ("Métier",       f"{mi[1]} {mi[0]}"),
                ("Demandé par",  src.get("demande_par","—")),
                ("Téléphone",    src.get("telephone","—")),
                ("N° Employé",   src.get("no_employe","—")),
                ("Pavillon",     src.get("pavillon","—")),
                ("Étage",        src.get("etage","—")),
                ("Aile",         src.get("aile","—")),
                ("Local",        src.get("local","—")),
                ("Localisation", src.get("localisation","—")),
            ]
            rows_html = "".join(
                f'<tr><td style="color:#8a9ab0;padding:3px 0;width:42%;font-size:0.76rem">{l}</td>'
                f'<td style="font-weight:500;color:{"#aaa" if v in ("—","NAN") else "#1a2a3a"};font-size:0.78rem">{v}</td></tr>'
                for l,v in champs
            )
            st.markdown(f"""
            <div style="background:white;border:1px solid #e8edf3;border-radius:9px;padding:0.9rem">
              <div style="display:flex;justify-content:space-between;margin-bottom:8px">
                <span style="font-family:'IBM Plex Mono',monospace;font-size:0.76rem;font-weight:600;color:#0D2B55">
                  {src.get('no_bon','— en cours —')}</span>
                <span class="badge" style="background:{pf};color:{pc}">{prio} — {pl}</span>
              </div>
              <table style="width:100%;border-collapse:collapse">{rows_html}</table>
              <div style="margin-top:7px;background:#f4f8fc;border-left:3px solid #006D77;border-radius:0 5px 5px 0;padding:7px 9px;font-size:0.78rem;color:#3a5068">
                <b>Problème :</b> {src.get('remarque_courte','—')}<br>
                <span style="font-size:0.73rem;color:#7a8a9a">{src.get('remarques','—')}</span>
              </div>
              {"<div style='margin-top:7px;background:#FFEBEE;border-radius:5px;padding:5px 9px;font-size:0.76rem;color:#C62828;font-weight:600'>🚨 URGENCE — Intervention immédiate</div>" if src.get('urgence') else ""}
            </div>""", unsafe_allow_html=True)
            if bon:
                st.download_button("⬇️ Télécharger BT (JSON)",
                                   data=json.dumps(bon,ensure_ascii=False,indent=2),
                                   file_name=f"{bon['no_bon']}.json",
                                   mime="application/json", use_container_width=True)

# ════════════════════════════════════════════════════════════════
# PAGE STATISTIQUES
# ════════════════════════════════════════════════════════════════
def page_statistiques(df_raw, gmao_log):
    st.markdown('<div class="section-title">📊 Statistiques — Analyse des requêtes HMR</div>', unsafe_allow_html=True)
    if df_raw.empty:
        st.info("Aucune donnée. Lance agentvoix pour recevoir des appels.")
        return

    total     = len(df_raw)
    trans     = sum(1 for n in df_raw["N°"] if n in gmao_log)
    urgences  = int(df_raw["_urgent"].sum())
    taux      = round(trans/total*100) if total else 0

    c1,c2,c3,c4 = st.columns(4)
    for col,(num,label,couleur) in zip([c1,c2,c3,c4],[
        (total,f"{taux}% transféré","#0D2B55"),
        (trans,"Transférées GMAO","#2E7D32"),
        (total-trans,"En attente","#E65100"),
        (urgences,"Urgences P1","#C62828"),
    ]):
        with col:
            st.markdown(f'<div class="kpi"><div class="kpi-n" style="color:{couleur};">{num}</div><div class="kpi-l">{label}</div></div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    c1,c2 = st.columns(2)
    with c1:
        st.markdown("**Requêtes par métier**")
        mc = df_raw["_metier"].value_counts().reset_index(); mc.columns=["Métier","N"]
        fig = px.bar(mc,x="N",y="Métier",orientation="h",color="N",color_continuous_scale="Blues",height=280)
        fig.update_layout(margin=dict(l=0,r=0,t=5,b=0),showlegend=False,coloraxis_showscale=False,plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        st.markdown("**Répartition par priorité**")
        pc = df_raw["_priorite"].value_counts().reset_index(); pc.columns=["P","N"]
        fig2 = px.pie(pc,values="N",names="P",color="P",
                      color_discrete_map={"P1":"#C62828","P2":"#E65100","P3":"#1565C0","P4":"#6A1B9A"},height=280)
        fig2.update_layout(margin=dict(l=0,r=0,t=5,b=0),paper_bgcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig2,use_container_width=True)

    c3,c4 = st.columns(2)
    with c3:
        st.markdown("**Statut GMAO**")
        sd = pd.DataFrame({"Statut":["Transférées","En attente"],"N":[trans,total-trans]})
        fig3 = px.bar(sd,x="Statut",y="N",color="Statut",
                      color_discrete_map={"Transférées":"#2E7D32","En attente":"#E65100"},height=230)
        fig3.update_layout(margin=dict(l=0,r=0,t=5,b=0),showlegend=False,plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig3,use_container_width=True)
    with c4:
        st.markdown("**Top pavillons**")
        pv = df_raw["Pavillon"].value_counts().head(6).reset_index(); pv.columns=["Pavillon","N"]
        fig4 = px.bar(pv,x="Pavillon",y="N",color="N",color_continuous_scale="Teal",height=230)
        fig4.update_layout(margin=dict(l=0,r=0,t=5,b=0),showlegend=False,coloraxis_showscale=False,plot_bgcolor="rgba(0,0,0,0)",paper_bgcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig4,use_container_width=True)

    st.markdown("**Tableau complet**")
    df_show = df_raw[["N°","Date d'ouverture","Récepteur","Demandé par","Pavillon","Remarque courte","_priorite"]].copy()
    df_show.columns = ["N°","Date","Récepteur","Demandé par","Pavillon","Remarque","Priorité"]
    st.dataframe(df_show, use_container_width=True, height=250)

    csv = df_raw[COLS_REGISTRE].to_csv(index=False, encoding="utf-8-sig")
    st.download_button("⬇️ Exporter CSV complet",
                       data=csv.encode("utf-8-sig"),
                       file_name=f"registre_HMR_{datetime.now().strftime('%Y%m%d')}.csv",
                       mime="text/csv")

# ════════════════════════════════════════════════════════════════
# PAGE DOCUMENTS
# ════════════════════════════════════════════════════════════════
def page_documents():
    st.markdown('<div class="section-title">🗂️ Documents — Fichiers du système</div>', unsafe_allow_html=True)
    tab1, tab2, tab3 = st.tabs(["📋 Registre GMAO", "📍 Locaux HMR", "🔧 Tâches standards"])

    with tab1:
        st.markdown("**Registre des appels de service**")
        existe = os.path.exists(FICHIER_REGISTRE)
        st.markdown(f"Fichier : `{FICHIER_REGISTRE}` — {'✅ Présent' if existe else '❌ Absent'}")
        if existe:
            df = charger_registre()
            st.info(f"{len(df)} bons de travail enregistrés.")
            if not df.empty:
                st.dataframe(df[COLS_REGISTRE].head(20), use_container_width=True, height=300)
        else:
            st.warning("Le fichier sera créé automatiquement lors du premier appel.")

    with tab2:
        st.markdown("**Référentiel des locaux HMR — 7 464 locaux**")
        existe = os.path.exists(FICHIER_LOCAUX)
        st.markdown(f"Fichier : `{FICHIER_LOCAUX}` — {'✅ Présent' if existe else '❌ Absent'}")
        if existe:
            df_loc = charger_locaux(200)
            st.info(f"Aperçu des 200 premiers locaux sur 7 464 au total.")
            filtre = st.text_input("🔎 Filtrer par numéro ou description")
            if filtre:
                df_loc = df_loc[df_loc["numero"].str.contains(filtre,case=False,na=False) |
                                df_loc["description"].str.contains(filtre,case=False,na=False)]
            st.dataframe(df_loc, use_container_width=True, height=350)
        else:
            st.warning(f"Placer `{FICHIER_LOCAUX}` dans le même dossier.")

    with tab3:
        st.markdown("**Tâches standards HMR — Intéral**")
        existe = os.path.exists(FICHIER_TACHES)
        st.markdown(f"Fichier : `{FICHIER_TACHES}` — {'✅ Présent' if existe else '❌ Absent'}")
        if existe:
            df_t = charger_taches()
            st.info(f"{len(df_t)} tâches HMR chargées.")
            f_cat = st.selectbox("Filtrer par catégorie",
                                 ["Toutes"] + sorted(df_t["categorie"].dropna().unique().tolist()))
            if f_cat != "Toutes":
                df_t = df_t[df_t["categorie"] == f_cat]
            st.dataframe(df_t[["no_tache","description","recepteur","ressource"]],
                         use_container_width=True, height=350)
        else:
            st.warning(f"Placer `{FICHIER_TACHES}` dans le même dossier.")

# ════════════════════════════════════════════════════════════════
# PAGE AGENT IA
# ════════════════════════════════════════════════════════════════
def page_agent_ia():
    st.markdown('<div class="section-title">🤖 Agent IA — Configuration OpenAI</div>', unsafe_allow_html=True)
    config = charger_json(FICHIER_CONFIG, {})

    tab1, tab2, tab3 = st.tabs(["🔑 Clés API", "🎙️ Voix et STT", "📝 Script agent"])

    with tab1:
        st.markdown("**Configuration Azure OpenAI / OpenAI**")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**OpenAI (pour démarrage rapide)**")
            openai_key = st.text_input("Clé API OpenAI", value=config.get("openai_key",""),
                                       type="password", placeholder="sk-...")
            openai_model = st.selectbox("Modèle", ["gpt-4o-mini","gpt-4o","gpt-3.5-turbo"],
                                        index=0)
            st.caption("⚠️ Données transmises aux USA — documenter dans l'EFVP (Loi 25)")

        with col2:
            st.markdown("**Azure OpenAI (recommandé production)**")
            azure_endpoint = st.text_input("Endpoint Azure", value=config.get("azure_endpoint",""),
                                           placeholder="https://votre-ressource.openai.azure.com")
            azure_key  = st.text_input("Clé Azure OpenAI", value=config.get("azure_key",""), type="password")
            azure_region = st.selectbox("Région", ["canadaeast","eastus","westeurope"], index=0)
            st.caption("✅ Région Canada East — conforme Loi 25")

        st.markdown("**Deepgram STT (transcription vocale)**")
        c1,c2 = st.columns(2)
        with c1:
            deepgram_key = st.text_input("Clé API Deepgram", value=config.get("deepgram_key",""),
                                         type="password", placeholder="Token Deepgram")
        with c2:
            deepgram_lang = st.selectbox("Langue", ["fr","fr-CA","en-US"], index=1)
            st.caption("✅ Serveurs disponibles au Canada")

        st.markdown("**Azure TTS (synthèse vocale)**")
        c1,c2 = st.columns(2)
        with c1:
            azure_tts_key = st.text_input("Clé Azure Speech", value=config.get("azure_tts_key",""), type="password")
        with c2:
            voix = st.selectbox("Voix française", ["fr-CA-SylvieNeural","fr-CA-AntoineNeural","fr-FR-DeniseNeural"], index=0)

        if st.button("💾 Sauvegarder configuration", type="primary"):
            config.update({
                "openai_key": openai_key, "openai_model": openai_model,
                "azure_endpoint": azure_endpoint, "azure_key": azure_key,
                "azure_region": azure_region, "deepgram_key": deepgram_key,
                "deepgram_lang": deepgram_lang, "azure_tts_key": azure_tts_key,
                "voix_tts": voix,
            })
            sauvegarder_json(FICHIER_CONFIG, config)
            st.success("✅ Configuration sauvegardée dans `agentvoix_config.json`")

        st.markdown("---")
        st.markdown("**Test de connexion OpenAI**")
        if st.button("🧪 Tester OpenAI"):
            key = config.get("openai_key","")
            if not key:
                st.error("Aucune clé OpenAI configurée.")
            else:
                try:
                    import urllib.request, json as _json
                    req = urllib.request.Request(
                        "https://api.openai.com/v1/chat/completions",
                        data=_json.dumps({"model":"gpt-4o-mini","messages":[{"role":"user","content":"Réponds juste OK"}],"max_tokens":5}).encode(),
                        headers={"Authorization":f"Bearer {key}","Content-Type":"application/json"},
                        method="POST"
                    )
                    with urllib.request.urlopen(req, timeout=10) as r:
                        res = _json.loads(r.read())
                        reponse = res["choices"][0]["message"]["content"]
                        st.success(f"✅ Connexion OpenAI OK — Réponse : {reponse}")
                except Exception as e:
                    st.error(f"❌ Erreur connexion : {e}")

    with tab2:
        st.markdown("**Pipeline vocal — latence cible < 1.5s**")
        etapes = [
            ("Twilio WebSocket","Réception audio en streaming","~0 ms","#2E7D32"),
            ("Deepgram Nova-2","Transcription STT français","~300 ms","#1565C0"),
            ("Azure OpenAI GPT-4o mini","Classification + réponse","~400 ms","#6A1B9A"),
            ("Azure Neural TTS","Synthèse vocale streamée","~200 ms","#E65100"),
            ("Twilio","Lecture audio à l'appelant","~100 ms","#37474F"),
        ]
        total_ms = 0
        for nom,desc,ms,couleur in etapes:
            val = int(ms.replace("~","").replace(" ms",""))
            total_ms += val
            st.markdown(f"""
            <div class="workflow-node">
              <div style="width:10px;height:10px;border-radius:50%;background:{couleur};flex-shrink:0"></div>
              <div style="flex:1"><b style="font-size:0.85rem">{nom}</b>
              <div style="font-size:0.75rem;color:#8a9ab0">{desc}</div></div>
              <div style="font-size:0.85rem;font-weight:600;color:{couleur}">{ms}</div>
            </div>""", unsafe_allow_html=True)
        st.markdown(f"""
        <div style="text-align:right;margin-top:8px;font-size:0.9rem">
          <b>Latence totale estimée : <span style="color:#2E7D32">~{total_ms} ms</span></b>
          &nbsp;✅ dans le budget
        </div>""", unsafe_allow_html=True)

    with tab3:
        st.markdown("**Script d'introduction — Loi 25 (obligatoire)**")
        st.code(
            "Bonjour ! Service de maintenance des bâtiments de l'Hôpital Maisonneuve-Rosemont.\n"
            "Je suis l'agent IA. Cet appel est traité par une intelligence artificielle\n"
            "et sera enregistré aux fins de traitement de votre demande uniquement.\n"
            "Les informations collectées sont sécurisées conformément à la Loi 25.\n"
            "Dites 'humain' ou appuyez sur le 0 en tout temps pour parler à une personne.",
            language=None
        )
        st.markdown("**Questions posées — dans l'ordre**")
        for i,(cle,question,aide) in enumerate(QUESTIONS_AGENT,1):
            st.markdown(f"**{i}.** {question}  \n`{cle}` — *{aide}*")

# ════════════════════════════════════════════════════════════════
# PAGE WORKFLOW N8N
# ════════════════════════════════════════════════════════════════
def page_workflow():
    st.markdown('<div class="section-title">🔄 Workflow n8n — Architecture complète</div>', unsafe_allow_html=True)

    tab1, tab2, tab3 = st.tabs(["📐 Architecture", "📋 Nœuds détaillés", "⚙️ Variables d'environnement"])

    with tab1:
        st.markdown("**Flux complet — de l'appel au bon de travail**")
        noeuds = [
            ("1","Webhook Twilio","Point d'entrée","Reçoit l'audio de chaque segment de parole via WebSocket","#1565C0","✅"),
            ("2","Valider appel","Nettoyage","Normalise l'entrée Twilio, détecte urgence immédiatement","#37474F","✅"),
            ("3","Switch Urgence","Décision","Branche 0=urgence (escalade immédiate), Branche 1=normal","#C62828","✅"),
            ("4","Escalade Urgence","Urgence","Transfert téléphonique vers permanence via API Twilio","#C62828","🔧"),
            ("5","Notifier Teams Urgence","Notification","Alerte instantanée dans le canal Teams urgences","#E65100","🔧"),
            ("6","Lire session","PostgreSQL","Récupère l'état de la conversation en cours","#6A1B9A","🔧"),
            ("7","Préparer dialogue","Logique métier","Décide quelle question poser, enregistre la réponse précédente","#6A1B9A","✅"),
            ("8","Switch Formulaire","Décision","Branche 0=complet (classifier), Branche 1=poser prochaine question","#37474F","✅"),
            ("9","Sauvegarder session","PostgreSQL","Persiste l'état en base pour reprendre si appel coupé","#6A1B9A","🔧"),
            ("10","TTS Question","Azure TTS","Convertit la prochaine question en audio français","#0277BD","🔧"),
            ("11","Réponse Twilio","TwiML","Envoie l'audio + écoute la réponse STT","#1565C0","🔧"),
            ("12","Classifier OpenAI","Azure OpenAI","GPT-4o mini classe la tâche Intéral depuis les remarques","#6A1B9A","🔧"),
            ("13","Parser résultat","Code","Assemble le bon de travail complet, génère le N° BT","#37474F","✅"),
            ("14","Écrire PostgreSQL","Base données","Insère la requête complète — Streamlit lit depuis ici","#2E7D32","🔧"),
            ("15","Écrire Excel","FastAPI local","Ajoute la ligne dans Requetes_Telephoniques_GMAO.xlsx","#2E7D32","🔧"),
            ("16","Notifier Teams","Microsoft Teams","Notification au coordonnateur avec lien vers Streamlit","#E65100","🔧"),
            ("17","Email récepteur","SMTP","Email à l'équipe concernée (plomberie, électricité…)","#E65100","🔧"),
            ("18","TTS Confirmation","Azure TTS","Confirmation vocale avec N° BT et équipe assignée","#0277BD","🔧"),
            ("19","Fin appel","TwiML","Joue la confirmation et raccroche proprement","#1565C0","🔧"),
            ("20","Nettoyer session","PostgreSQL","Supprime la session temporaire après BT créé","#6A1B9A","🔧"),
        ]
        for no,nom,type_,desc,couleur,statut in noeuds:
            st.markdown(f"""
            <div class="workflow-node">
              <div style="width:24px;height:24px;border-radius:50%;background:{couleur};
                color:white;font-size:0.7rem;font-weight:600;display:flex;
                align-items:center;justify-content:center;flex-shrink:0">{no}</div>
              <div style="flex:1">
                <div style="font-size:0.85rem;font-weight:600;color:#0D2B55">{nom}
                  <span style="font-size:0.68rem;color:#8a9ab0;font-weight:400;margin-left:6px">{type_}</span>
                </div>
                <div style="font-size:0.75rem;color:#5a7080">{desc}</div>
              </div>
              <div style="font-size:0.8rem">{statut}</div>
            </div>""", unsafe_allow_html=True)
        st.caption("✅ = Logique prête | 🔧 = Nécessite configuration des clés API")

    with tab2:
        st.markdown("**Configuration requise par nœud**")
        configs = {
            "Twilio": ["TWILIO_ACCOUNT_SID","TWILIO_AUTH_TOKEN","TEL_PERMANENCE"],
            "Azure OpenAI": ["AZURE_OPENAI_ENDPOINT","AZURE_OPENAI_KEY","AZURE_REGION"],
            "Azure TTS": ["AZURE_TTS_KEY","AZURE_REGION","VOIX_TTS"],
            "Deepgram STT": ["DEEPGRAM_API_KEY","DEEPGRAM_LANGUAGE"],
            "Microsoft Teams": ["TEAMS_TEAM_ID","TEAMS_CHANNEL_URGENCES","TEAMS_CHANNEL_REQUETES"],
            "PostgreSQL": ["PG_HOST","PG_PORT","PG_DB","PG_USER","PG_PASSWORD"],
            "n8n": ["N8N_HOST","N8N_PORT","STREAMLIT_HOST"],
        }
        for service, vars_ in configs.items():
            with st.expander(f"🔧 {service}"):
                for v in vars_:
                    st.code(v)

        st.markdown("**Schéma PostgreSQL**")
        st.code("""
-- Table principale des requêtes
CREATE TABLE requetes_gmao (
    id               SERIAL PRIMARY KEY,
    no_bon           VARCHAR(30) UNIQUE,
    date_ouverture   TIMESTAMP,
    recepteur        VARCHAR(30),
    emetteur         VARCHAR(20) DEFAULT 'Agent IA',
    no_employe       VARCHAR(20),
    demande_par      VARCHAR(100),
    demande_pour     VARCHAR(100),
    local            VARCHAR(30),
    site             VARCHAR(10) DEFAULT 'HMR',
    pavillon         VARCHAR(50),
    etage            VARCHAR(10),
    aile             VARCHAR(10),
    localisation     VARCHAR(100),
    telephone        VARCHAR(20),
    remarque_courte  TEXT,
    remarques        TEXT,
    tache            VARCHAR(100),
    metier           VARCHAR(50),
    priorite         VARCHAR(5),
    urgence          BOOLEAN DEFAULT FALSE,
    transfert_gmao   BOOLEAN DEFAULT FALSE,
    transfert_le     TIMESTAMP,
    transfert_par    VARCHAR(50),
    session_id       VARCHAR(50),
    created_at       TIMESTAMP DEFAULT NOW()
);

-- Sessions de conversation en cours
CREATE TABLE agentvoix_sessions (
    session_id   VARCHAR(50) PRIMARY KEY,
    session_data JSONB,
    created_at   TIMESTAMP DEFAULT NOW()
);
        """, language="sql")

    with tab3:
        st.markdown("**Variables d'environnement n8n — à configurer dans Settings → Variables**")
        vars_env = {
            "TWILIO_ACCOUNT_SID":       "ACxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx",
            "TWILIO_AUTH_TOKEN":        "votre_token_twilio",
            "TEL_PERMANENCE":           "+15141234567",
            "AZURE_OPENAI_ENDPOINT":    "https://votre-ressource.openai.azure.com",
            "AZURE_OPENAI_KEY":         "votre_cle_azure_openai",
            "AZURE_REGION":             "canadaeast",
            "AZURE_TTS_KEY":            "votre_cle_azure_speech",
            "DEEPGRAM_API_KEY":         "votre_cle_deepgram",
            "TEAMS_TEAM_ID":            "id_equipe_teams",
            "TEAMS_CHANNEL_URGENCES":   "id_canal_urgences",
            "TEAMS_CHANNEL_REQUETES":   "id_canal_requetes",
            "N8N_HOST":                 "votre-domaine-n8n.com",
            "STREAMLIT_HOST":           "votre-domaine-streamlit.com",
            "PG_HOST":                  "localhost",
            "PG_PORT":                  "5432",
            "PG_DB":                    "agentvoix",
            "PG_USER":                  "agentvoix_user",
            "PG_PASSWORD":              "votre_mot_de_passe",
        }
        df_vars = pd.DataFrame([{"Variable":k,"Valeur exemple":v} for k,v in vars_env.items()])
        st.dataframe(df_vars, use_container_width=True, hide_index=True)
        csv = df_vars.to_csv(index=False)
        st.download_button("⬇️ Télécharger template variables", data=csv,
                           file_name="n8n_variables_env.csv", mime="text/csv")

# ════════════════════════════════════════════════════════════════
# PAGE CONFIGURATION
# ════════════════════════════════════════════════════════════════
def page_configuration():
    st.markdown('<div class="section-title">⚙️ Configuration — Système AgentVoix</div>', unsafe_allow_html=True)
    tab1, tab2, tab3, tab4 = st.tabs(["📡 Récepteurs", "🗺️ Routage", "🏢 Sites", "❓ Questions agent"])

    with tab1:
        st.markdown("**Récepteurs actifs — HMR**")
        rows = [{"Code":k,"Métier":v[0],"Icône":v[1]} for k,v in RECEPTEURS.items()]
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    with tab2:
        st.markdown("**Table de routage automatique — mots-clés → récepteur**")
        routing = [
            ("fuite, eau, plomberie, toilette, bouché, drain, robinet, lavabo, douche, vapeur", "HMR-CEPLOM", "P2"),
            ("lumière, électr, prise, panne, interrupteur, ampoule, fluorescent, courant, circuit", "HMR-CEELE", "P2"),
            ("température, ventilation, chauffage, thermostat, humidité, climatisation, froid", "HMR-CEMMF", "P2"),
            ("serrure, porte, clé, verrou, poignée, cadenas, barre de panique, anti-fugue", "HMR-CESER", "P3"),
            ("frigo, congélateur, réfrigérateur, machine à glace, climatiseur, fontaine réfrigérée", "HMR-CEFRIG", "P2"),
            ("lit, civière, chariot, roulette, fauteuil, lève-personne, ridelle, manivelle", "HMR-CEMAINT", "P3"),
            ("meuble, plancher, plafond, tuile, fenêtre, mur, store, rideau, peinture, vitre", "HMR-CEMENU", "P4"),
            ("lavage, nettoyage, désinfection, poubelle, déchets, salubrité, punaises, souris", "HMR-CESALU", "P3"),
            ("urgence, feu, incendie, danger, blessé, inondation, gaz, électrocution, critique", "HMR-CEMAINT", "P1 🚨"),
        ]
        df_r = pd.DataFrame(routing, columns=["Mots-clés détectés","Récepteur","Priorité"])
        st.dataframe(df_r, use_container_width=True, hide_index=True, height=320)
        st.info("En production : routage effectué par Azure OpenAI GPT-4o mini sur les 156 tâches standards HMR.")

    with tab3:
        st.markdown("**Sites du CIUSSS de l'Est-de-l'Île-de-Montréal**")
        sites = [
            ("HMR",   "Hôpital Maisonneuve-Rosemont",                    "Principal","✅ Actif"),
            ("IUSMM", "Institut Universitaire Santé Mentale de Montréal","Secondaire","🔧 À configurer"),
            ("HSCO",  "Hôpital Santa Cabrini Ospedale",                  "Secondaire","🔧 À configurer"),
            ("PDI",   "Point-de-Île",                                    "Secondaire","🔧 À configurer"),
            ("LTEAS", "Lucille-Teasdale",                                "Secondaire","🔧 À configurer"),
            ("PMCS",  "Polonais Marie-Curie-Sklodowaska",                "Secondaire","🔧 À configurer"),
            ("SLSM",  "Saint-Léonard et Saint-Michel",                  "Secondaire","🔧 À configurer"),
            ("BOSCO", "Boscoville",                                       "Secondaire","🔧 À configurer"),
        ]
        df_s = pd.DataFrame(sites, columns=["Code","Nom complet","Type","Statut"])
        st.dataframe(df_s, use_container_width=True, hide_index=True)

    with tab4:
        st.markdown("**Séquence des questions — Agent IA**")
        for i,(cle,question,aide) in enumerate(QUESTIONS_AGENT,1):
            col_n,col_q,col_o = st.columns([1,8,3])
            with col_n: st.markdown(f"**{i}**")
            with col_q: st.markdown(f"**{question}**  \n*{aide}*")
            with col_o:
                obl = cle not in ("aile","localisation")
                st.markdown(f"{'🔴 Obligatoire' if obl else '🟡 Optionnel (NAN)'}")

# ════════════════════════════════════════════════════════════════
# PAGE UTILISATEURS
# ════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════
# PAGE ELEVENLABS
# ════════════════════════════════════════════════════════════════
def page_elevenlabs():
    st.markdown('<div class="section-title">🎙️ ElevenLabs — Agent vocal IA</div>', unsafe_allow_html=True)
    config = charger_json(FICHIER_CONFIG, {})

    tab1, tab2, tab3 = st.tabs(["🔑 Configuration", "🧪 Test agent", "📊 Monitoring conversations"])

    # ── TAB 1 : Configuration ──────────────────────────────────
    with tab1:
        st.markdown("**Connexion ElevenLabs**")
        c1, c2 = st.columns(2)
        with c1:
            el_key = st.text_input(
                "Clé API ElevenLabs",
                value=config.get("elevenlabs_key",""),
                type="password",
                placeholder="sk_xxxxxxxxxxxxxxxxxxxxxxxx",
                help="Paramètres → API Keys dans ElevenLabs"
            )
            el_agent_id = st.text_input(
                "Agent ID",
                value=config.get("elevenlabs_agent_id",""),
                placeholder="agent_xxxxxxxxxxxxxxxx",
                help="Agents → ton agent HMR → ID"
            )
        with c2:
            el_voice_id = st.text_input(
                "Voice ID (optionnel)",
                value=config.get("elevenlabs_voice_id",""),
                placeholder="Voice ID ElevenLabs",
                help="Voix → ID de la voix française choisie"
            )
            el_webhook = st.text_input(
                "URL Webhook n8n",
                value=config.get("elevenlabs_webhook",""),
                placeholder="https://votre-n8n.com/webhook/agentvoix-hmr",
                help="URL du webhook n8n qui reçoit les données après chaque appel"
            )

        st.markdown("---")
        st.markdown("**Prompt système de l'agent HMR**")
        prompt_defaut = """Tu es l'agent IA de réception des appels de service technique de l'Hôpital Maisonneuve-Rosemont, CIUSSS de l'Est-de-l'Île-de-Montréal.

INTRODUCTION OBLIGATOIRE (Loi 25) :
"Bonjour ! Service de maintenance des bâtiments de l'Hôpital Maisonneuve-Rosemont. Je suis l'agent IA. Cet appel est traité par une intelligence artificielle et sera enregistré aux fins de traitement de votre demande uniquement, conformément à la Loi 25. Dites 'humain' en tout temps pour parler à une personne."

QUESTIONS À POSER UNE PAR UNE (attendre la réponse avant de continuer) :
1. Quel est votre prénom et nom ?
2. Votre numéro de téléphone ou extension ?
3. Votre numéro d'employé ? (ex: HMR-8301)
4. Cette demande est pour vous-même ou une autre personne ?
5. Dans quel pavillon se trouve le problème ? (Maisonneuve, Lavoisier, Guy-Bernier, Rachel-Tourigny)
6. À quel étage ?
7. Dans quelle aile ? (dites 'je ne sais pas' si inconnu)
8. Quel est le numéro du local ou de la salle ?
9. Précisez la localisation exacte. (optionnel)
10. En quelques mots, quel est le problème ?
11. Décrivez le problème en détail.

RÈGLES IMPORTANTES :
- Si la personne dit 'je ne sais pas', 'inconnu', 'pas certain' → inscrire NAN et passer à la question suivante
- Si tu détectes les mots urgence, feu, incendie, danger, blessé → dire immédiatement "Je vous transfère d'urgence vers la permanence" et terminer l'appel
- Ne jamais sauter une question
- Parler en français québécois naturel
- À la fin : confirmer le numéro de bon de travail généré

CONCLUSION :
"Merci [nom]. Votre demande est enregistrée. L'équipe technique va intervenir. Bonne journée !" """

        prompt = st.text_area(
            "Script de l'agent",
            value=config.get("elevenlabs_prompt", prompt_defaut),
            height=350,
            help="Ce prompt définit le comportement de l'agent lors des appels"
        )

        st.markdown("---")
        st.markdown("**Configuration webhook — Structure JSON envoyée à n8n**")
        st.code("""{
  "event": "conversation_completed",
  "conversation_id": "conv_xxx",
  "agent_id": "agent_xxx",
  "data": {
    "demande_par":     "Marie Tremblay",
    "telephone":       "4381234567",
    "no_employe":      "HMR-8301",
    "demande_pour":    "Marie Tremblay",
    "pavillon":        "Maisonneuve",
    "etage":           "5",
    "aile":            "A",
    "local":           "MA05102",
    "localisation":    "Bureau infirmières",
    "remarque_courte": "Lumière brisée",
    "remarques":       "Les tubes fluorescents sont grillés",
    "urgence":         false,
    "duree_appel":     180
  }
}""", language="json")

        if st.button("💾 Sauvegarder configuration ElevenLabs", type="primary"):
            config.update({
                "elevenlabs_key":      el_key,
                "elevenlabs_agent_id": el_agent_id,
                "elevenlabs_voice_id": el_voice_id,
                "elevenlabs_webhook":  el_webhook,
                "elevenlabs_prompt":   prompt,
            })
            sauvegarder_json(FICHIER_CONFIG, config)
            st.success("✅ Configuration ElevenLabs sauvegardée !")

    # ── TAB 2 : Test agent ──────────────────────────────────────
    with tab2:
        st.markdown("**Tester la connexion ElevenLabs**")
        key = config.get("elevenlabs_key","")
        agent_id = config.get("elevenlabs_agent_id","")

        if not key:
            st.warning("⚠️ Configure d'abord ta clé API dans l'onglet Configuration.")
        else:
            c1, c2, c3 = st.columns(3)
            with c1:
                if st.button("🔗 Tester connexion API", use_container_width=True):
                    try:
                        import urllib.request, json as _json
                        req = urllib.request.Request(
                            "https://api.elevenlabs.io/v1/user",
                            headers={"xi-api-key": key},
                            method="GET"
                        )
                        with urllib.request.urlopen(req, timeout=10) as r:
                            data = _json.loads(r.read())
                            st.success(f"✅ Connexion OK — Compte : {data.get('first_name','')} {data.get('last_name','')}")
                            st.info(f"Caractères restants : {data.get('subscription',{}).get('character_limit',0) - data.get('subscription',{}).get('character_count',0):,}")
                    except Exception as e:
                        st.error(f"❌ Erreur : {e}")
            with c2:
                if st.button("🎙️ Lister les voix FR", use_container_width=True):
                    try:
                        import urllib.request, json as _json
                        req = urllib.request.Request(
                            "https://api.elevenlabs.io/v1/voices",
                            headers={"xi-api-key": key},
                            method="GET"
                        )
                        with urllib.request.urlopen(req, timeout=10) as r:
                            data = _json.loads(r.read())
                            voix_fr = [v for v in data.get("voices",[])
                                       if "french" in str(v.get("labels",{})).lower()
                                       or "fr" in str(v.get("labels",{})).lower()]
                            if voix_fr:
                                for v in voix_fr[:5]:
                                    st.markdown(f"**{v['name']}** — `{v['voice_id']}`")
                            else:
                                st.info("Aucune voix française trouvée. Toutes les voix disponibles :")
                                for v in data.get("voices",[])[:5]:
                                    st.markdown(f"**{v['name']}** — `{v['voice_id']}`")
                    except Exception as e:
                        st.error(f"❌ Erreur : {e}")
            with c3:
                if st.button("🤖 Vérifier agent", use_container_width=True):
                    if not agent_id:
                        st.error("Agent ID non configuré.")
                    else:
                        try:
                            import urllib.request, json as _json
                            req = urllib.request.Request(
                                f"https://api.elevenlabs.io/v1/convai/agents/{agent_id}",
                                headers={"xi-api-key": key},
                                method="GET"
                            )
                            with urllib.request.urlopen(req, timeout=10) as r:
                                data = _json.loads(r.read())
                                st.success(f"✅ Agent trouvé : **{data.get('name','—')}**")
                                st.json({
                                    "name": data.get("name"),
                                    "language": data.get("conversation_config",{}).get("agent",{}).get("language","—"),
                                    "status": "Actif"
                                })
                        except Exception as e:
                            st.error(f"❌ Erreur : {e}")

        st.markdown("---")
        st.markdown("**Générer un audio de test**")
        texte_test = st.text_input(
            "Texte à synthétiser",
            value="Bonjour, je suis l'agent IA de l'Hôpital Maisonneuve-Rosemont. Comment puis-je vous aider ?",
        )
        voice_id_test = config.get("elevenlabs_voice_id","") or "21m00Tcm4TlvDq8ikWAM"

        if st.button("🔊 Générer audio test", use_container_width=True):
            if not key:
                st.error("Clé API requise.")
            else:
                try:
                    import urllib.request, json as _json
                    body = _json.dumps({
                        "text": texte_test,
                        "model_id": "eleven_multilingual_v2",
                        "voice_settings": {"stability": 0.5, "similarity_boost": 0.75}
                    }).encode()
                    req = urllib.request.Request(
                        f"https://api.elevenlabs.io/v1/text-to-speech/{voice_id_test}",
                        data=body,
                        headers={
                            "xi-api-key": key,
                            "Content-Type": "application/json",
                            "Accept": "audio/mpeg"
                        },
                        method="POST"
                    )
                    with urllib.request.urlopen(req, timeout=15) as r:
                        audio = r.read()
                        st.audio(audio, format="audio/mp3")
                        st.success("✅ Audio généré avec succès !")
                except Exception as e:
                    st.error(f"❌ Erreur génération audio : {e}")

    # ── TAB 3 : Monitoring ─────────────────────────────────────
    with tab3:
        st.markdown("**Conversations récentes — Agent HMR**")
        key = config.get("elevenlabs_key","")
        agent_id = config.get("elevenlabs_agent_id","")

        if not key or not agent_id:
            st.warning("⚠️ Configure la clé API et l'Agent ID dans l'onglet Configuration.")
        else:
            col_refresh, _ = st.columns([2,6])
            with col_refresh:
                if st.button("🔄 Charger conversations", use_container_width=True):
                    try:
                        import urllib.request, json as _json
                        req = urllib.request.Request(
                            f"https://api.elevenlabs.io/v1/convai/conversations?agent_id={agent_id}&page_size=20",
                            headers={"xi-api-key": key},
                            method="GET"
                        )
                        with urllib.request.urlopen(req, timeout=10) as r:
                            data = _json.loads(r.read())
                            conversations = data.get("conversations", [])
                            st.session_state["el_conversations"] = conversations
                    except Exception as e:
                        st.error(f"❌ Erreur : {e}")

            convs = st.session_state.get("el_conversations", [])
            if not convs:
                st.info("Clique sur 'Charger conversations' pour voir les appels récents.")
            else:
                st.markdown(f"**{len(convs)} conversation(s) chargée(s)**")
                for conv in convs:
                    status = conv.get("status","—")
                    duree  = conv.get("metadata",{}).get("duration_secs",0)
                    conv_id = conv.get("conversation_id","—")
                    date_ts = conv.get("metadata",{}).get("start_time_unix_secs",0)
                    date_str = datetime.fromtimestamp(date_ts).strftime("%Y-%m-%d %H:%M") if date_ts else "—"

                    couleur = "#2E7D32" if status == "done" else "#E65100"
                    icone   = "✅" if status == "done" else "🔄"

                    st.markdown(f"""
                    <div class="req">
                      <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap">
                        <span class="req-no">{conv_id[:20]}...</span>
                        <span class="badge" style="background:{'#e8f5e9' if status=='done' else '#fff3e0'};
                          color:{couleur}">{icone} {status}</span>
                        <span style="font-size:0.78rem;color:#6B7A8D">⏱ {duree}s</span>
                        <span style="margin-left:auto;font-size:0.72rem;color:#aaa">{date_str}</span>
                      </div>
                    </div>""", unsafe_allow_html=True)

                # Stats rapides
                st.markdown("---")
                st.markdown("**Statistiques**")
                total_conv = len(convs)
                terminees  = sum(1 for c in convs if c.get("status") == "done")
                duree_moy  = sum(c.get("metadata",{}).get("duration_secs",0) for c in convs) / max(total_conv,1)

                c1,c2,c3 = st.columns(3)
                with c1:
                    st.markdown(f'<div class="kpi"><div class="kpi-n" style="color:#003A70">{total_conv}</div><div class="kpi-l">Conversations</div></div>', unsafe_allow_html=True)
                with c2:
                    st.markdown(f'<div class="kpi"><div class="kpi-n" style="color:#2E7D32">{terminees}</div><div class="kpi-l">Complétées</div></div>', unsafe_allow_html=True)
                with c3:
                    st.markdown(f'<div class="kpi"><div class="kpi-n" style="color:#00A1DE">{duree_moy:.0f}s</div><div class="kpi-l">Durée moyenne</div></div>', unsafe_allow_html=True)


def page_utilisateurs():
    st.markdown('<div class="section-title">👥 Utilisateurs — Gestion des accès</div>', unsafe_allow_html=True)
    users = charger_json(FICHIER_USERS, USERS_DEFAUT)

    col1, col2 = st.columns([3,2])
    with col1:
        st.markdown("**Comptes actifs**")
        rows = [{"Identifiant":k,"Nom":v["nom"],"Rôle":v["role"].capitalize()}
                for k,v in users.items()]
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    with col2:
        st.markdown("**Rôles et permissions**")
        roles = [
            ("admin",        "Toutes les pages + gestion utilisateurs"),
            ("coordonnateur","Accueil, Requêtes, Stats, Documents"),
            ("technicien",   "Accueil, Requêtes (lecture)"),
        ]
        for role, perms in roles:
            st.markdown(f"""
            <div class="config-card">
              <div class="config-title">👤 {role.capitalize()}</div>
              <div style="font-size:0.78rem;color:#5a7080">{perms}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("**Ajouter un utilisateur**")
    with st.form("add_user"):
        c1,c2,c3,c4 = st.columns(4)
        with c1: new_id  = st.text_input("Identifiant")
        with c2: new_nom = st.text_input("Nom complet")
        with c3: new_role= st.selectbox("Rôle",["coordonnateur","technicien","admin"])
        with c4: new_mdp = st.text_input("Mot de passe", type="password")
        submitted = st.form_submit_button("➕ Ajouter", type="primary")
        if submitted:
            if new_id and new_nom and new_mdp:
                if new_id in users:
                    st.error(f"L'identifiant '{new_id}' existe déjà.")
                else:
                    users[new_id] = {"mdp":hash_mdp(new_mdp),"role":new_role,"nom":new_nom}
                    sauvegarder_json(FICHIER_USERS, users)
                    st.success(f"✅ Utilisateur '{new_id}' créé.")
                    st.rerun()
            else:
                st.error("Tous les champs sont obligatoires.")

    st.markdown("**Changer un mot de passe**")
    with st.form("change_mdp"):
        c1,c2,c3 = st.columns(3)
        with c1: chg_id  = st.selectbox("Utilisateur", list(users.keys()))
        with c2: chg_mdp = st.text_input("Nouveau mot de passe", type="password")
        with c3: st.markdown("<br>", unsafe_allow_html=True)
        if st.form_submit_button("🔑 Changer"):
            if chg_mdp:
                users[chg_id]["mdp"] = hash_mdp(chg_mdp)
                sauvegarder_json(FICHIER_USERS, users)
                st.success(f"✅ Mot de passe de '{chg_id}' mis à jour.")
            else:
                st.error("Entrez un nouveau mot de passe.")

# ════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════
def main():
    init_auth()

    if not st.session_state.logged_in:
        page_login()
        return

    # Charger données
    df_raw   = charger_registre()
    gmao_log = charger_json(FICHIER_GMAO_LOG, {})
    if not df_raw.empty:
        df_raw["_transfere"] = df_raw["N°"].apply(lambda n: n in gmao_log)

    # Sidebar
    page, f_statut, f_prio, f_metier, f_search = render_sidebar(df_raw, gmao_log)

    # Header
    st.markdown(f"""
    <div class="top-bar">
      <div>
        <h1>🏥 AgentVoix — Hôpital Maisonneuve-Rosemont</h1>
        <div class="sub">CIUSSS de l'Est-de-l'Île-de-Montréal · Service technique · Maintenance des bâtiments</div>
      </div>
      <div class="right">
        {datetime.now().strftime('%A %d %B %Y')}<br>{datetime.now().strftime('%H:%M')}
      </div>
    </div>
    """, unsafe_allow_html=True)

    # Routing pages
    if   "Accueil"       in page: page_accueil(df_raw, gmao_log)
    elif "Requêtes"      in page: page_requetes(df_raw, gmao_log, f_statut, f_prio, f_metier, f_search)
    elif "Simulation"    in page: page_simulation()
    elif "Statistiques"  in page: page_statistiques(df_raw, gmao_log)
    elif "Documents"     in page: page_documents()
    elif "Agent IA"      in page: page_agent_ia()
    elif "Workflow"      in page: page_workflow()
    elif "Configuration" in page: page_configuration()
    elif "ElevenLabs"    in page: page_elevenlabs()
    elif "Utilisateurs"  in page: page_utilisateurs()

if __name__ == "__main__":
    main()
